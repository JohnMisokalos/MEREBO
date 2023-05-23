from sklearn.preprocessing import LabelEncoder
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
import unicodedata as ud

#functions to check whether the ML model approves the match

def svmapproves(model, lb, comp):
    arr1q = list(comp[0].toarray())
    arr1 = np.array(arr1q)
    arr2q = list(comp[1].toarray())
    arr2 = np.array(arr2q)
    finalarray = np.concatenate([arr1[0], arr2[0]])

    y_test = model.predict([finalarray])
    label = lb.inverse_transform(y_test)
    return label

def svmapproveslist(model, lb, m1, m2):
    for i in range(m1.shape[0]):

        arr1 = list(m1.toarray())
        # arr1 = np.array(arr1q)
        arr2 = list(m2.toarray())
        # arr2 = np.array(arr2q)
        finalarray= []
        for j in range(0, len(arr1)):
            finalarray.append(np.concatenate([arr1[j], arr2[j]]))
        finarray = np.array(finalarray)

        y_test = model.predict(finarray)
        label = lb.inverse_transform(y_test)
        # print(label, len(label))
        return label

percentageoftraining = 80

full_dataset = openpyxl.load_workbook("K:/Thesis/datasets/total.xlsx")
train = openpyxl.load_workbook("K:/Thesis/datasets/altbl2/artsmall262/artificialsetsmalltrain.xlsx")

sh = full_dataset.active
all_items =[]

#optimal variables found for this example
cutoff = 0.5
decrease = 0.9

for i in range(2, sh.max_row+1):

    cell_obj = sh.cell(row=i, column=1)
    if(cell_obj.value!= 0):
        x = cell_obj.value.lower()
        d = {ord('\N{COMBINING ACUTE ACCENT}'):None}
        y = ud.normalize('NFD',x).translate(d)

        bow = y.split(" ")
        all_items.append(bow)

text_data = [" ".join(i) for i in all_items]

tfidf = TfidfVectorizer(binary=True)
vector_matrix = tfidf.fit_transform(text_data)

sh = train.active

train_items =[]
tags = []

for i in range(1, sh.max_row+1):
    cell_obj = sh.cell(row=i, column=3)
    if (cell_obj.value == None) or (cell_obj.value == 0):
        bow = ""
    else:
        x = cell_obj.value.lower()
        d = {ord('\N{COMBINING ACUTE ACCENT}'):None}
        y = ud.normalize('NFD',x).translate(d)
        bow = y.split(" ")
    train_items.append(bow)

    cell_obj = sh.cell(row=i, column=6)

    if (cell_obj.value == None) or (cell_obj.value == 0) :
        bow = ""
    else:
        x = cell_obj.value.lower()
        d = {ord('\N{COMBINING ACUTE ACCENT}'):None}
        y = ud.normalize('NFD',x).translate(d)
        bow = y.split(" ")
    train_items.append(bow)


    cell_obj = sh.cell(row=i, column=7).value
    tags.append(cell_obj)
train_data = [" ".join(i) for i in train_items]

train_matrix = tfidf.transform(train_data)

trainarr = list(train_matrix.toarray())

target = tags
lb = LabelEncoder() # to produce the distinct labels.
Y = lb.fit_transform(target) #Fit label encoder and return encoded labels.



trainarrayfinal = []
for i in range(0, len(trainarr)):
    if(i%2==0):

        trainarrayfinal.append(np.concatenate([trainarr[i], trainarr[i+1]]))

trayfinal = np.array(trainarrayfinal)

teynarrayfinal = []


model = RandomForestClassifier(criterion = 'entropy', bootstrap = 'true')


model.fit(trayfinal, Y)

# sh = full_dataset.active

most_similar = []


for i in range (0, vector_matrix.shape[0]):
    print(i)
    similarity = cosine_similarity(vector_matrix[i], vector_matrix)

    supername = sh.cell(row=i+1, column=2).value
    id = sh.cell(row=i+1, column=1).value
    relsim = 0
    compval = [0,0]
    mostsim1 = [0,0,0,0,0,0]
    mostsim2 = [0,0,0,0,0,0]
    pos = [0,0]

    a = vector_matrix.copy()
    for z in range(vector_matrix.shape[0]):
        a[z] = vector_matrix[i]

    flag = svmapproveslist(model, lb, a, vector_matrix)
    # print(flag)
    for z in range(0, len(flag)):
        if flag[z] == 'F':
            similarity[0][z] = decrease*similarity[0][z]

    for j in range(1, sh.max_row+1):
        #flag = svmapproves(model, lb, [test_matrix[i],test_matrix[j-2]])
        # print(j)
        if (supername == 1) and (sh.cell(row=j, column=2).value == 2) and (similarity[0][j-1]>compval[0]) and (similarity[0][j-1]>cutoff):
            compval[0] = similarity[0][j-1]
            mostsim1 = [sh.cell(row=j, column=1).value, sh.cell(row=j, column=2).value, sh.cell(row=j, column=3).value, sh.cell(row=j, column=4).value,  sh.cell(row=j, column=5).value, compval[0]]
            pos[0] = j
        elif (supername == 1) and (sh.cell(row=j, column=2).value == 3) and (similarity[0][j-1]>compval[1]) and (similarity[0][j-1]>cutoff):
            compval[1] = similarity[0][j-1]
            mostsim2 = [sh.cell(row=j, column=1).value, sh.cell(row=j, column=2).value, sh.cell(row=j, column=3).value, sh.cell(row=j, column=4).value,  sh.cell(row=j, column=5).value, compval[1]]
            pos[1] = j
        elif (supername == 2) and (sh.cell(row=j, column=2).value == 1) and (similarity[0][j-1]>compval[0]) and (similarity[0][j-1]>cutoff):
            compval[0] = similarity[0][j-1]
            mostsim1 = [sh.cell(row=j, column=1).value, sh.cell(row=j, column=2).value, sh.cell(row=j, column=3).value, sh.cell(row=j, column=4).value,  sh.cell(row=j, column=5).value, compval[0]]
            pos[0] = j
        elif (supername == 2) and (sh.cell(row=j, column=2).value == 3) and (similarity[0][j-1]>compval[1]) and (similarity[0][j-1]>cutoff):
            compval[1] = similarity[0][j-1]
            mostsim2 = [sh.cell(row=j, column=1).value, sh.cell(row=j, column=2).value, sh.cell(row=j, column=3).value, sh.cell(row=j, column=4).value,  sh.cell(row=j, column=5).value, compval[1]]
            pos[1] = j
        elif (supername == 3) and (sh.cell(row=j, column=2).value == 1) and (similarity[0][j-1]>compval[0]) and (similarity[0][j-1]>cutoff):
            compval[0] = similarity[0][j-1]
            mostsim1 = [sh.cell(row=j, column=1).value, sh.cell(row=j, column=2).value, sh.cell(row=j, column=3).value, sh.cell(row=j, column=4).value,  sh.cell(row=j, column=5).value, compval[0]]
            pos[0] = j
        elif (supername == 3) and (sh.cell(row=j, column=2).value == 2) and (similarity[0][j-1]>compval[1]) and (similarity[0][j-1]>cutoff):
            compval[1] = similarity[0][j-1]
            mostsim2 = [sh.cell(row=j, column=1).value, sh.cell(row=j, column=2).value, sh.cell(row=j, column=3).value, sh.cell(row=j, column=4).value,  sh.cell(row=j, column=5).value, compval[1]]
            pos[1] = j

    relsim = cosine_similarity(vector_matrix[pos[0]], vector_matrix[pos[1]])

    x = [id, supername, sh.cell(row=i+1, column=3).value, mostsim1[0], mostsim1[1], mostsim1[2], mostsim1[3], mostsim1[4], mostsim1[5], mostsim2[0], mostsim2[1], mostsim2[2], mostsim2[3], mostsim2[4], mostsim2[5], relsim]
    print(x)
    most_similar.append(x)

df = pd.DataFrame(most_similar)


df.to_excel('K:/Thesis/datasets/solrup.xlsx', header=False, index=False)
